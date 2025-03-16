# tests/test_save_email_to_drafts.py
import os
import base64
import tempfile
import json
import time
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pytest
from googleapiclient.errors import HttpError
from unittest.mock import MagicMock, patch

from src.utils import save_email_to_drafts as sed

# --- Dummy Classes and Functions for Patching ---

class DummyCredentials:
    valid = True
    expired = False
    refresh_token = True

class DummyRequest:
    def execute(self):
        return {"id": "draft_id", "message": {"id": "msg_id"}}

class DummyDrafts:
    def create(self, userId, body):
        return DummyRequest()

class DummyService:
    def users(self):
        return self
    def drafts(self):
        return DummyDrafts()

def dummy_build(serviceName, version, credentials):
    return DummyService()

# --- Tests ---

def test_create_message():
    """
    Test that create_message returns a correctly encoded email message.
    """
    sender = "sender@example.com"
    recipient = "recipient@example.com"
    subject = "Test Subject"
    message_text = "<p>Test email content</p>"
    message_dict = sed.create_message(sender, recipient, subject, message_text)
    assert "raw" in message_dict
    raw = message_dict["raw"]
    # Decode the raw message back to text.
    decoded = base64.urlsafe_b64decode(raw.encode()).decode()
    decoded_lower = decoded.lower()
    # Check that the content-type header is present.
    assert "content-type: text/html" in decoded_lower
    # Check headers in a case-insensitive manner.
    assert f"subject: {subject.lower()}" in decoded_lower
    assert f"to: {recipient.lower()}" in decoded_lower
    assert f"from: {sender.lower()}" in decoded_lower
    assert "test email content" in decoded_lower

@patch('src.utils.save_email_to_drafts.get_credentials')
@patch('src.utils.save_email_to_drafts.build')
def test_save_email_to_drafts_success(mock_build, mock_get_credentials):
    """
    Test save_email_to_drafts using the legacy plain-text parameters.
    """
    mock_get_credentials.return_value = DummyCredentials()
    mock_build.return_value = DummyService()
    # Call using keyword arguments.
    draft_id = sed.save_email_to_drafts(sender='test@example.com',
                                        recipient='recipient@example.com',
                                        subject='Subject',
                                        body='Body')
    assert draft_id == "draft_id"

@patch('src.utils.save_email_to_drafts.get_credentials')
@patch('src.utils.save_email_to_drafts.build')
def test_save_email_to_drafts_http_error(mock_build, mock_get_credentials):
    """
    Test that when the Gmail API call raises an HttpError, the function returns None.
    """
    mock_get_credentials.return_value = DummyCredentials()
    mock_build.return_value = DummyService()
    # Create a dummy response for HttpError.
    dummy_resp = MagicMock()
    dummy_resp.reason = "Bad Request"
    dummy_resp.status = 400
    dummy_resp.headers = {}
    # Patch the execute method of DummyRequest to raise HttpError.
    with patch.object(DummyRequest, "execute", side_effect=HttpError(resp=dummy_resp, content=b'Error')):
         draft_id = sed.save_email_to_drafts(sender='test@example.com',
                                             recipient='recipient@example.com',
                                             subject='Subject',
                                             body='Body')
    assert draft_id is None

@patch('src.utils.save_email_to_drafts.get_credentials', return_value=None)
def test_save_email_to_drafts_no_credentials(mock_get_credentials):
    """
    Test that if no credentials are returned, save_email_to_drafts returns None.
    """
    draft_id = sed.save_email_to_drafts(sender='test@example.com',
                                        recipient='recipient@example.com',
                                        subject='Subject',
                                        body='Body')
    assert draft_id is None

def test_save_email_to_drafts_plain():
    """
    Test save_email_to_drafts legacy usage with keyword arguments.
    """
    with patch('src.utils.save_email_to_drafts.get_credentials', return_value=DummyCredentials()), \
         patch('src.utils.save_email_to_drafts.build', return_value=DummyService()):
        draft_id = sed.save_email_to_drafts(sender='test@example.com',
                                            recipient='recipient@example.com',
                                            subject='Subject',
                                            body='Body')
        assert draft_id == "draft_id"

def test_save_email_to_drafts_mime():
    """
    Test save_email_to_drafts when a MIME message is provided.
    """
    sender = "sender@example.com"
    recipient = "recipient@example.com"
    subject = "Test Subject"
    body = "<p>Test email content with inline images</p>"
    # Create a simple MIME message.
    mime_message = MIMEMultipart('related')
    mime_message['From'] = sender
    mime_message['To'] = recipient
    mime_message['Subject'] = subject
    mime_message.attach(MIMEText(body, 'html'))
    with patch('src.utils.save_email_to_drafts.get_credentials', return_value=DummyCredentials()), \
         patch('src.utils.save_email_to_drafts.build', return_value=DummyService()):
        draft_id = sed.save_email_to_drafts(mime_message=mime_message)
        assert draft_id == "draft_id"

# --- Tests for save_data_to_excel ---

def test_save_data_to_excel_new_file(tmp_path):
    """
    Test that save_data_to_excel writes data to a new Excel file correctly.
    """
    file_path = tmp_path / "test_output.xlsx"
    data = {
        'saving_file_time': "2025/03/16 12:00",
        'company': 'Test Company',
        'website': 'http://test.com',
        'main_business': 'Test business',
        'cooperation_letter_conter': 'Test letter',
        'recipient_email': 'recipient@example.com',
        'contact_person': 'Test Contact'
    }
    sed.save_data_to_excel(data, str(file_path))
    assert file_path.exists()
    workbook = openpyxl.load_workbook(str(file_path))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    for key in data.keys():
        assert key in headers
    row_data = [cell.value for cell in sheet[2]]
    company_index = headers.index('company')
    assert row_data[company_index] == 'Test Company'

def test_save_data_to_excel_existing_file(tmp_path):
    """
    Test that save_data_to_excel appends data to an existing Excel file.
    """
    file_path = tmp_path / "test_output.xlsx"
    # Create a file with headers.
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ['saving_file_time', 'company', 'website', 'main_business',
               'cooperation_letter_conter', 'recipient_email', 'contact_person']
    sheet.append(headers)
    workbook.save(str(file_path))
    data = {
        'saving_file_time': "2025/03/16 12:00",
        'company': 'Test Company Existing',
        'website': 'http://test.com',
        'main_business': 'Test business',
        'cooperation_letter_conter': 'Test letter',
        'recipient_email': 'recipient@example.com',
        'contact_person': 'Test Contact'
    }
    sed.save_data_to_excel(data, str(file_path))
    workbook = openpyxl.load_workbook(str(file_path))
    sheet = workbook.active
    # Expect two rows: header and one data row.
    assert sheet.max_row == 2
    row_data = [cell.value for cell in sheet[2]]
    company_index = headers.index('company')
    assert row_data[company_index] == 'Test Company Existing'

def test_save_data_to_excel_exception(tmp_path, monkeypatch):
    """
    Test that save_data_to_excel logs an error when an exception occurs.
    This test forces an exception by monkeypatching os.path.exists.
    """
    file_path = tmp_path / "test_output.xlsx"
    monkeypatch.setattr(os.path, "exists", lambda x: False)
    data = {
        'saving_file_time': "2025/03/16 12:00",
        'company': 'Test Company',
        'website': 'http://test.com',
        'main_business': 'Test business',
        'cooperation_letter_conter': 'Test letter',
        'recipient_email': 'recipient@example.com',
        'contact_person': 'Test Contact'
    }
    # The call should not raise an exception (it logs an error instead).
    sed.save_data_to_excel(data, str(file_path))
